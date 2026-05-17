"""
Export endpoints for secretary:
  GET /export/{race_id}/excel  — protocol as .xlsx
  GET /export/{race_id}/pdf    — protocol as .pdf
"""
import io
import re
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from app.db.database import get_db
from app.models.incident import ProtocolEntry
from app.models.race import Race

router = APIRouter(prefix="/export", tags=["export"])


@router.get("/{race_id}/excel")
async def export_excel(race_id: int, db: AsyncSession = Depends(get_db)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    race = await db.get(Race, race_id)
    if not race:
        raise HTTPException(status_code=404, detail="Race not found")

    entries_result = await db.execute(
        select(ProtocolEntry)
        .where(ProtocolEntry.race_id == race_id)
        .order_by(ProtocolEntry.sequence_number)
    )
    entries = entries_result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Protocol"

    # Header row
    headers = ["#", "Pilot(s)", "Violation", "Decision", "Penalty", "Post", "Marshal", "Judge", "Time"]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, entry in enumerate(entries, start=2):
        violation_text = entry.transcript_raw or entry.violation_type or ""
        ws.cell(row=row_idx, column=1, value=entry.sequence_number)
        ws.cell(row=row_idx, column=2, value=entry.pilot_numbers)
        ws.cell(row=row_idx, column=3, value=violation_text)
        ws.cell(row=row_idx, column=4, value=entry.decision_type)
        ws.cell(row=row_idx, column=5, value=entry.penalty_detail or "")
        ws.cell(row=row_idx, column=6, value=entry.post_label)
        ws.cell(row=row_idx, column=7, value=entry.marshal_name)
        ws.cell(row=row_idx, column=8, value=entry.judge_name)
        ws.cell(row=row_idx, column=9, value=entry.created_at.strftime("%H:%M:%S"))

    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = re.sub(r"[^\w]", "_", race.name, flags=re.ASCII)
    filename = f"protocol_{safe_name}_{datetime.now():%Y%m%d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{race_id}/pdf")
async def export_pdf(race_id: int, db: AsyncSession = Depends(get_db)):
    import os
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    fonts_dir = os.path.join(os.path.dirname(__file__), "..", "..", "fonts")
    pdfmetrics.registerFont(TTFont("DejaVu", os.path.join(fonts_dir, "DejaVuSans.ttf")))
    pdfmetrics.registerFont(TTFont("DejaVu-Bold", os.path.join(fonts_dir, "DejaVuSans-Bold.ttf")))

    race = await db.get(Race, race_id)
    if not race:
        raise HTTPException(status_code=404, detail="Race not found")

    entries_result = await db.execute(
        select(ProtocolEntry)
        .where(ProtocolEntry.race_id == race_id)
        .order_by(ProtocolEntry.sequence_number)
    )
    entries = entries_result.scalars().all()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        topMargin=20 * mm,
        bottomMargin=15 * mm,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
    )

    title_style = ParagraphStyle("title", fontName="DejaVu-Bold", fontSize=16, spaceAfter=6)
    sub_style = ParagraphStyle("sub", fontName="DejaVu", fontSize=9, textColor=colors.HexColor("#555555"), spaceAfter=10)

    headers = ["#", "Пилот(ы)", "Нарушение", "Решение", "Штраф", "Пост", "Маршал", "Судья", "Время"]
    # landscape A4 = 297mm, margins 15+15 = 267mm available
    col_widths = [10, 22, 40, 30, 32, 23, 43, 43, 22]  # sum = 265mm

    table_data = [headers]
    for entry in entries:
        violation_text = entry.transcript_raw or entry.violation_type or ""
        table_data.append([
            str(entry.sequence_number),
            entry.pilot_numbers or "",
            violation_text,
            entry.decision_type or "",
            entry.penalty_detail or "",
            entry.post_label or "",
            entry.marshal_name or "",
            entry.judge_name or "",
            entry.created_at.strftime("%H:%M:%S"),
        ])

    table = Table(table_data, colWidths=[w * mm for w in col_widths], repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "DejaVu-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "DejaVu"),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EAF2FF")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]))

    title = Paragraph(f"Протокол гонки: {race.name}", title_style)
    subtitle = Paragraph(
        f"{race.venue} · {datetime.now().strftime('%d.%m.%Y')} · {len(entries)} записей",
        sub_style,
    )
    doc.build([title, subtitle, table])
    buf.seek(0)

    safe_name = re.sub(r"[^\w]", "_", race.name, flags=re.ASCII)
    filename = f"protocol_{safe_name}_{datetime.now():%Y%m%d}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
